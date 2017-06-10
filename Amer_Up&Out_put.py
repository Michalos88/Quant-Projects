from math import exp
import numpy as np
import openpyxl as op


# Initialize parameters
K = 100
T = 1
S = 100
r = 0.05
H = 110
X_rebate = 1
N = 4
dx = 0.2524

wb = op.load_workbook('Euro_prices_F5.3.xlsx')


def call(x,y):
    sheet_call = wb.get_sheet_by_name('Call')
    c = sheet_call.cell(row=x, column=y).value
    return c


def put(x,y):
    sheet_put = wb.get_sheet_by_name('Put')
    c = sheet_put.cell(row=x, column=y).value
    return c


# Precompute constants

dt = T/N
edx = exp(dx)
infl = exp(r*dt)

# Initialize the matrices

St = np.zeros((2*N + 1, N + 1), dtype=np.float64)  # Asset prices
Q = np.zeros((2*N + 1, N + 1), dtype=np.float64)  # State prices
pu = np.zeros((2*N + 1, N + 1), dtype=np.float64)
pm = np.zeros((2*N + 1, N + 1), dtype=np.float64)
pd = np.zeros((2*N + 1, N + 1), dtype=np.float64)

# Initialize asset prices
St[2*N, N] = S*exp(-N*dx)
for j in range(2*N-1,-1,-1):
    St[j,N] = St[j+1,N]*edx


# For each time step:

for i in range(0,N+1):
    for j in range(N-i,N+1,1):
        sum = 0
        for k in range(j-2,N,1): #Not sure
            sum = sum + Q[k, i]*(St[k, N] - St[j + 1,N])
        # k = j-1
        # sum = sum + Q[k, i] * (St[k, N] - St[j + 1, N])

        D = St[j,N] - St[j+1,N]
        Q[j,i] = (call(j+1, i+1) - sum)/D

    for j in range(N+i, N, -1):
        # print(j, i)
        sum = 0
        for k in range(N+i, j+2, 1):  # Not sure
            if k<=8:
                sum = sum + Q[k, i] * (St[j-1, N] - St[k, N])

        D = St[j-1, N] - St[j , N]
        Q[j, i] = (put(j - 1, i + 1) - sum) / D


# Implied Trinomial Tree Transition Probabilities

for i in range(0,N):
    for j in range(N-i,N+1,1):
        if j == N - i:
            pu[N - i, i] = infl * Q[N - i - 1, i + 1] / Q[N - i, i]
            pm[j, i] = (infl * St[j, N] - St[j + 1, N] - pu[j, i] * (St[j - 1, N] - St[j + 1, N])) / (St[j, N] - St[j + 1, N])
            pd[j, i] = 1 - pm[j, i] - pu[j, i]

        elif j == N - i + 1:
            pu[N - i + 1, i] = (infl * Q[N - i, i + 1] - pm[N - i, i] * Q[N - i, i]) / Q[N - i + 1, i]
            pm[j, i] = (infl * St[j, N] - St[j + 1, N] - pu[j, i] * (St[j - 1, N] - St[j + 1, N])) / (St[j, N] - St[j + 1, N])
            pd[j, i] = 1 - pm[j, i] - pu[j, i]
        else:
            pu[j, i] = (infl * Q[j - 1, i + 1] - pd[j - 2, i] * Q[j - 2, i] - pm[j - 1, i] * Q[j - 1, i]) / Q[j, i]
            pm[j, i] = (infl * St[j, N] - St[j + 1, N] - pu[j, i] * (St[j - 1, N] - St[j + 1, N])) / (St[j, N] - St[j + 1, N])
            pd[j, i] = 1 - pm[j, i] - pu[j, i]

    for j in range(N + i, N, -1):
        if j == N+i:
            pd[N+i,i] = infl*Q[N+i+1,i+1]/Q[N+i,i]
            pm[j, i] = (infl * St[j, N] - St[j - 1, N] - pd[j, i] * (St[j + 1, N] - St[j - 1, N])) / (St[j, N] - St[j - 1, N])
            pu[j, i] = 1 - pm[j, i] - pd[j, i]

        elif j == N + i - 1:
            pd[N + i - 1,i] = (infl*Q[N+i,i+1] - pm[N+i,i]*Q[N+i,i])/Q[N+i-1,i]
            pm[j, i] = (infl * St[j, N] - St[j - 1, N] - pd[j, i] * (St[j + 1, N] - St[j - 1, N])) / (St[j, N] - St[j - 1, N])
            pu[j, i] = 1 - pm[j, i] - pd[j, i]
        else:
            pd[j,i] = (infl * Q[j + 1, i + 1] - pu[j + 2, i] * Q[j + 2, i] - pm[j + 1, i] * Q[j + 1, i]) / Q[j, i]
            pm[j,i] = (infl * St[j, N] - St[j - 1, N] - pd[j, i] * (St[j + 1, N] - St[j - 1, N])) / (St[j, N] - St[j - 1, N])
            pu[j,i] = 1 - pm[j, i] - pd[j, i]

#Valution of American Up And Out Put Option

# Precompute constants
disc = exp(-r*dt)

# Initialize the matrix

C = np.zeros((2*N + 1, N + 1), dtype=np.float64)  # Option Value

# Initialize option values at maturity

for j in range(0,2*N+1):
    if St[j,N] < H:
        C[j,N] = max(0.0,K - St[j,N])
    else:
        C[j,N] = 0.0

for i in range(N-1,-1,-1):
    for j in range(N-i-1,N+i+1):
        if St[j,N] < H:
            C[j,i] = disc*(pu[j,i]*C[j-1,i+1] + pm[j,i]*C[j,i+1] + pd[j,i]*C[j+1,i+1])

            # Apply the early exercise condition
            C[j,i] = max(C[j,i],K - St[j,N])
        else:
            C[j,i] = X_rebate

# np.savetxt("AmericanUpnOutPut.csv", C, delimiter=",")

