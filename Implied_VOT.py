#  Only using toolbox functions for normal CDF calculation and mathematical tools and excel operations
from math import *
import random
from scipy.stats import norm
import openpyxl

# Opening the data file
wb = openpyxl.load_workbook('Data1.xlsx')
sheet1 = wb.get_sheet_by_name('SPY1118')

# Initializing constants:
T = 52/365  # Time to maturity = Recorded on 9/26
S = 214.95  # Underlying asset price
r = 0.4  # interest rate
tol = 10**-6  # tolerance
A = 2  # starting row in excel

# Initializing arrays for variables
K = [0]*101  # Array of strike prices
Vol = [0]*101  # Array of implied volatility
Bid = [0]*101
Ask = [0]*101
C = [0]*101

option = 'Call'

def impliedvol(S, K, T, r, option,C,tol):

    def CBS(S, K, T, r, sigma, option):
        # Calculations for the solution to BSM equation
        dplus = (1 / (sigma * sqrt(T))) * ((log(S / K)) + (r + (sigma ** 2) / 2) * T)
        dminus = (1 / (sigma * sqrt(T))) * ((log(S / K)) + (r - (sigma ** 2) / 2) * T)

        # Calculating price of Call and Put
        if option == 'Call':
            return S * norm.cdf(dplus) - K * exp(-r * T) * norm.cdf(dminus)
        elif option == 'Put':
            return K * exp(-r * T) * norm.cdf(-dminus) - S * norm.cdf(-dplus)


    def f(x, option):
        return CBS(S, K, T, r, x, option) - C


    # Calculating implied votality by bisection method
    t = 0
    while t < 1000:
        t = t + 1
        a = random.uniform(0, 100)
        b = random.uniform(0, 100)
        if (f(a, option)) * (f(b, option)) < 0:
            break
        else:
            continue



    def bsecvot(S, K, T, r, option, a, b):
        c = (a + b) / 2
        if f(a,option) == 0:
            return a

        elif f(b,option) == 0:
            return b

        elif f(a,option) * f(b,option) < 0:
            while abs(b - a) > tol:

                if f(a,option) * f(c,option) < 0:
                    b = c
                    c = (a + b) / 2
                   # print('C1')
                    continue

                elif f(b,option)*f(c,option) < 0:
                    a = c
                    c = (a + b) / 2
                    #print('C2')
                    continue
        return c

    return bsecvot(S, K, T, r, option, a, b)

while A < 124:
    A=A+1
    Bid[A] = sheet1.cell(row=A, column=4).value
    Ask[A] = sheet1.cell(row=A, column=5).value
    K[A] = sheet1.cell(row=A, column=1).value

    C[A] = (Bid[A] + Ask[A]) / 2
    Vol[A] = impliedvol(S,K[A],T,r,option,C[A],tol)

print(Vol)


